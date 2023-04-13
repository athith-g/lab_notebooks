from openpyxl import load_workbook

# Load Excel sheet
print("Opening file and loading contents...")
wb = load_workbook(filename='spreadsheets/c2combined.xlsx')
sheet = wb.active

ID = input("Enter ID Column Letter: ")
t = input("Enter Time Column Letter: ")
eo = input("Enter EMAOccasion Column Letter: ")
ec = input("Enter EMACompliance Letter: ")

i, j = 2, 2
while sheet[f"{ID}{i}"].value is not None:

    occ_i = (sheet[f"{ID}{i}"].value, sheet[f"{eo}{i}"].value)
    if not occ_i[1] or occ_i[1] == 0:
        i += 1
        j += 1
        continue

    # Get window of rows that have same ID and EMAOccasion
    occ_j = (sheet[f"{ID}{j}"].value, sheet[f"{eo}{j}"].value)
    while occ_i == occ_j:
        if i != j:
            _cell = sheet.cell(j, 21)
            _cell.number_format = 'General'
            sheet[f"{ec}{j}"] = ""

        j += 1
        occ_j = (sheet[f"{ID}{j}"].value, sheet[f"{eo}{j}"].value)

    i = j

# Save file
print("Saving File...")
wb.save(filename="spreadsheets/c2combined_test.xlsx")