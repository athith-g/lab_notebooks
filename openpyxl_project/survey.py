from openpyxl import load_workbook

# Load Excel sheet
print("Opening file and loading contents...")
wb = load_workbook(filename='spreadsheets2/reference.xlsx')
sheet = wb.active
ids = set()

i = 2
while sheet[i][0].value is not None:
    ids.add(sheet[i][0].value)
    i += 1

print(ids)

print("Opening file and loading contents...")
wb = load_workbook(filename='spreadsheets2/Timeline.xlsx')
sheet = wb.active

i = 1
date_codes = {}

while i < 23:
    if sheet[i][0].value == 'C':
        i += 1
        continue

    date_codes[sheet[i][0].value] = sheet[i][1].value
    i += 1


wb = load_workbook(filename='spreadsheets2/C1B1 Weekly Survey B_toupload.xlsx')
sheet = wb.active

i = 2
j = 2

for ID in ids:
    print("ID:", ID)
    while sheet[i][0].value == sheet[j][0].value:
        j += 1
        if sheet[i][0].value is None and sheet[j][0].value is None:
            break

    init_i = i
    if sheet[i][0].value != ID:
        print("Mismatch at i = ", i, "ID = ", ID)
        for date in date_codes:
            if date_codes[date] is None:
                continue
            sheet.insert_rows(i)
            sheet[i][0].value = ID
            sheet[i][1].value = (date_codes[date])
            sheet[i][2].value = date
            sheet[i][3].value = -999
            sheet[i][4].value = -999
            sheet[i][5].value = -999
            i += 1

    else:
        for date in date_codes:
            if date_codes[date] is None:
                continue
            if sheet[i][2].value == date:
                print("HIT:", sheet[i][2].value, date, ID)
                sheet[i][1].value = (date_codes[date])
                sheet[i][3].value = -999 if sheet[i][3].value is None else sheet[i][3].value
                sheet[i][4].value = -999 if sheet[i][4].value is None else sheet[i][4].value
                sheet[i][5].value = -999 if sheet[i][5].value is None else sheet[i][5].value
                i += 1
            else:
                print("MISS:", sheet[i][2].value, date, ID)
                sheet.insert_rows(i)
                sheet[i][0].value = ID
                sheet[i][1].value = (date_codes[date])
                sheet[i][2].value = date
                sheet[i][3].value = -999
                sheet[i][4].value = -999
                sheet[i][5].value = -999
                i += 1
                print(i)

    j = i


print("Saving File...")
wb.save(filename="spreadsheets2/t3.xlsx")