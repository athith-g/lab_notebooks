from openpyxl import load_workbook

# Load Excel sheet
wb = load_workbook(filename='spreadsheets/c2db2.xlsx')
sheet = wb.active

ID = input("Enter ID Column Letter: ")
t = input("Enter Time Column Letter: ")
eo = input("Enter EMAOccasion Column Letter: ")
ec = input("Enter EMACompliance Letter: ")

i, j = 2, 2
while sheet[f"{eo}{i}"].value is not None:
    # Skip row if EMAOccasion is 0 or EMACompliance is already calculated
    occ_i = (sheet[f"{ID}{i}"].value, sheet[f"{eo}{i}"].value)
    done = sheet[f"{ec}{i}"].value
    if not occ_i[1] or occ_i[1] == 0 or done is not None:
        i += 1
        j += 1
        continue

    # Get window of rows that have same ID and EMAOccasion
    occ_j = (sheet[f"{ID}{j}"].value, sheet[f"{eo}{j}"].value)
    while occ_i == occ_j:
        j += 1
        occ_j = (sheet[f"{ID}{j}"].value, sheet[f"{eo}{j}"].value)

    # Add Time values in window to set (only accepts unique values)
    nums = sheet[f"{eo}{i}"].value.split('-')
    s = set()
    for x in range(i, j):
        if sheet[f"{t}{x}"].value:
            if sheet[f"{t}{x}"].value in range(int(nums[0]), int(nums[1]) + 1):
                s.add(sheet[f"{t}{x}"].value)

    # Set compliance based on how many unique Time values there were (e.g., 0-> 0, 1-> 0.25, 2->0.5, 3->0.75, 4->1)
    comp = 0
    if len(s) == 1:
        comp = 0.25
    elif len(s) == 2:
        comp = 0.5
    elif len(s) == 3:
        comp = 0.75
    elif len(s) >= 4:
        comp = 1

    # Update compliance values
    for x in range(i, j):
        _cell = sheet.cell(x, 21)
        sheet[f"{ec}{x}"] = None
        _cell.number_format = 'Number'
        sheet[f"{ec}{x}"] = comp

    # Move on to next window
    i = j

# Save file
wb.save(filename="spreadsheets/c2db2_filled.xlsx")